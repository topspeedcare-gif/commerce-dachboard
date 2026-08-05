"""
dashboard/ad_report_importer.py — 쿠팡 윙 광고관리센터 "기간별 키워드" 리포트를
daily_metrics에 반영한다.

쿠팡은 셀러용 광고 성과 공개 API를 제공하지 않는다 (확인 완료) — 대신 광고관리센터에서
수동으로 받을 수 있는 엑셀을 등록하면, sync_job.py가 이미 채워둔 매출/판매량 위에
광고 데이터만 얹어서 calc_daily_metric()으로 다시 계산한다. 원가·수수료는 기존 설정값을
그대로 쓴다 — 이 파일이 새로 정의하는 값은 없다.
"""
from __future__ import annotations

import hashlib
from collections import defaultdict
from pathlib import Path

from dashboard.calc import calc_daily_metric
from shared.db import get_conn, upsert_daily_metric, get_setting, set_setting

REQUIRED_COLUMNS = ["날짜", "광고집행 옵션ID", "노출수", "클릭수", "광고비"]


def _num(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _fmt_date(raw) -> str:
    text = str(int(raw)) if isinstance(raw, float) else str(raw)
    text = text.strip()
    return f"{text[:4]}-{text[4:6]}-{text[6:8]}"


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _load_workbook_rows(path: Path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}, []
    header = [str(h or "").strip() for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    return idx, rows[1:]


def import_ad_report(path: str | Path) -> str:
    path = Path(path)
    if not path.exists():
        return f"❌ 파일을 찾지 못했습니다: {path}"

    digest = _sha256(path)
    setting_key = f"ad_report_imported:{digest}"
    if get_setting(setting_key):
        return f"ℹ️ 이미 등록된 파일입니다 (내용 동일): {path.name}\n중복 집계를 막기 위해 다시 반영하지 않았습니다."

    idx, data_rows = _load_workbook_rows(path)
    missing = [k for k in REQUIRED_COLUMNS if k not in idx]
    if missing:
        return f"❌ 엑셀 형식이 예상과 다릅니다 (쿠팡 윙 광고관리센터 '기간별 키워드' 리포트가 맞는지 확인하세요). 없는 컬럼: {missing}"

    agg: dict[tuple[str, str], dict] = defaultdict(lambda: {"impressions": 0, "clicks": 0, "spend": 0})
    for r in data_rows:
        if r[idx["날짜"]] is None:
            continue
        option_id_raw = r[idx["광고집행 옵션ID"]]
        if not option_id_raw:
            continue
        d = _fmt_date(r[idx["날짜"]])
        sku = str(int(_num(option_id_raw)))
        key = (d, sku)
        agg[key]["impressions"] += int(_num(r[idx["노출수"]]))
        agg[key]["clicks"] += int(_num(r[idx["클릭수"]]))
        agg[key]["spend"] += int(_num(r[idx["광고비"]]))

    if not agg:
        return "⚠️ 엑셀에서 유효한 데이터를 찾지 못했습니다."

    updated = 0
    unmatched = 0
    with get_conn() as conn:
        for (d, sku), ad in agg.items():
            row = conn.execute(
                "SELECT sales_qty, revenue FROM daily_metrics WHERE date = ? AND sku = ?", (d, sku)
            ).fetchone()
            if row is None:
                unmatched += 1
                sales_qty, revenue = 0, 0
            else:
                sales_qty, revenue = row["sales_qty"], row["revenue"]

            unit_cost = int(get_setting(f"unit_cost:{sku}", "0") or 0)
            metric = calc_daily_metric(
                date=d, sku=sku,
                sales_qty=sales_qty, revenue=revenue,
                ad_impressions=ad["impressions"], ad_clicks=ad["clicks"], ad_spend=ad["spend"],
                unit_cost=unit_cost,
            )
            upsert_daily_metric(metric)
            updated += 1

    set_setting(setting_key, path.name)
    total_spend = sum(a["spend"] for a in agg.values())
    dates = sorted({d for d, _ in agg})
    lines = [
        f"✅ 광고 리포트 등록 완료: {path.name}",
        f"  · 대상 기간: {dates[0]} ~ {dates[-1]}",
        f"  · 반영 SKU-일자 조합: {updated}건",
        f"  · 총 광고비: {total_spend:,}원",
    ]
    if unmatched:
        lines.append(f"  ⚠️ 그 중 {unmatched}건은 해당 날짜에 주문 데이터가 없어 매출 0으로 기록했습니다.")
    return "\n".join(lines)
